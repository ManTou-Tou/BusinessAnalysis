"""Excel 导入任务（Phase 9，queue=sync）。

经 Phase 3 run_lifecycle。_run 自管 SessionLocal，逐块「数据 + 进度」同事务提交
（svc.update_import_progress）：
- shops/products：upsert，从头重跑幂等；
- reviews：insert-only，从已提交 processed_rows **断点续跑**（防重复）；
- orders：按 order_no 边界分组，订单头 upsert + 明细先删后插；非相邻同 order_no → 整文件失败。
成功后删上传文件；失败/重试中保留（reviews 续跑、其它重跑需要它）。
"""
from __future__ import annotations

import os

from openpyxl import load_workbook

from app.core.celery_app import celery_app
from app.core.config import settings
from app.core.db import SessionLocal
from app.services import agent_task as svc
from app.services.imports import common as c
from app.services.imports import orders as orders_imp
from app.services.imports import products as products_imp
from app.services.imports import reviews as reviews_imp
from app.services.imports import shops as shops_imp
from app.tasks.base import run_lifecycle

_ROW_ENTITIES = {
    "shops": shops_imp,
    "products": products_imp,
    "reviews": reviews_imp,
}


class _Aborted(Exception):
    """处理中任务已非 running（被 reaper/cancel 接管）→ 中止，不再往非 running 任务落库。"""


def _is_blank(row) -> bool:
    return all(v is None for v in row)


def _prescan_orders(path: str) -> None:
    """orders：在任何写入前全量校验同 order_no 行相邻（非相邻 → 整文件失败，零脏写）。

    仅流式读 order_no 一列，维护 seen-closed 集合（O(distinct orders) 字符串，远小于全量
    行数据）——这是「整文件失败先于任何 chunk 落库」的必要全局状态（chunk 提交后再发现
    无法回滚已提交订单）。校验通过后正式处理才开始写。
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        rows_iter = wb.worksheets[0].iter_rows(values_only=True)
        header = next(rows_iter, None)
        if header is None:
            raise ValueError("empty sheet (no header row)")
        idx = c.map_header(header).get("order_no")
        if idx is None:
            raise ValueError("missing required columns: ['order_no']")
        seen: set[str] = set()
        current: str | None = None
        for row in rows_iter:
            if _is_blank(row):
                continue
            val = row[idx] if idx < len(row) else None
            if val is None or not str(val).strip():
                continue  # 空 order_no 行在正式处理时记为行错误，这里不影响相邻性
            key = str(val).strip()
            if key != current:
                if key in seen:
                    raise ValueError(
                        f"non-adjacent duplicate order_no {key!r}: rows of the same "
                        "order must be contiguous"
                    )
                if current is not None:
                    seen.add(current)
                current = key
    finally:
        wb.close()


def _progress(processed, total, inserted, updated, err: c.RowErrorCollector) -> dict:
    return {
        "processed_rows": processed,
        "total_rows": total,
        "inserted": inserted,
        "updated": updated,
        "error_count": err.count,
        "errors": err.errors,
    }


def _commit_progress(db, task_id, prog) -> None:
    # 进度落在 output_json.progress（reviews 续跑从此读 processed_rows；与 _run/_success 一致）。
    if not svc.update_import_progress(db, task_id, {"progress": prog}):
        raise _Aborted()


def _drive_rows(db, task_id, rows_iter, mapping, mod, account_id, shop_id, err, resume_offset, inserted, updated):
    chunk_size = settings.bulk_insert_chunk_size
    processed = resume_offset
    data_row_no = 0
    buffer: list[tuple[int, tuple]] = []

    def flush() -> None:
        nonlocal inserted, updated, processed, buffer
        ins, upd, errs = mod.process_chunk(db, account_id, shop_id, mapping, buffer)
        for rn, msg in errs:
            err.add(rn, msg)
        inserted += ins
        updated += upd
        processed = buffer[-1][0]
        _commit_progress(db, task_id, _progress(processed, None, inserted, updated, err))
        buffer = []

    for row in rows_iter:
        data_row_no += 1
        if data_row_no <= resume_offset or _is_blank(row):
            continue
        buffer.append((data_row_no, row))
        if len(buffer) >= chunk_size:
            flush()
    if buffer:
        flush()
    return _progress(processed, data_row_no, inserted, updated, err)


def _drive_orders(db, task_id, rows_iter, mapping, shop_id, err, inserted, updated):
    chunk_budget = settings.bulk_insert_chunk_size
    processed = 0
    data_row_no = 0
    seen_closed: set[str] = set()
    current_no: str | None = None
    current_rows: list[tuple[int, tuple]] = []
    rows_since_commit = 0

    def finalize_group() -> None:
        nonlocal inserted, updated, processed, rows_since_commit
        ins, upd, errs = orders_imp.flush_order(db, shop_id, current_no, current_rows, mapping)
        for rn, msg in errs:
            err.add(rn, msg)
        inserted += ins
        updated += upd
        processed = current_rows[-1][0]
        rows_since_commit += len(current_rows)
        seen_closed.add(current_no)

    for row in rows_iter:
        data_row_no += 1
        if _is_blank(row):
            continue
        try:
            order_no = c.req_str(c.raw(row, mapping, "order_no"), "order_no")
        except ValueError as exc:
            err.add(data_row_no, str(exc))
            continue
        if order_no != current_no:
            if current_no is not None:
                finalize_group()
                if rows_since_commit >= chunk_budget:
                    _commit_progress(
                        db, task_id, _progress(processed, None, inserted, updated, err)
                    )
                    rows_since_commit = 0
                current_rows = []
            # 防御性二次校验（_prescan_orders 已在写入前全量保证相邻；此处理论不触发）。
            if order_no in seen_closed:
                raise ValueError(
                    f"non-adjacent duplicate order_no {order_no!r}: rows of the same "
                    "order must be contiguous"
                )
            current_no = order_no
        current_rows.append((data_row_no, row))

    if current_no is not None and current_rows:
        finalize_group()
    _commit_progress(db, task_id, _progress(processed, data_row_no, inserted, updated, err))
    return _progress(processed, data_row_no, inserted, updated, err)


def _run(agent_task_id: int) -> dict:
    db = SessionLocal()
    try:
        task = svc.get_by_id(db, agent_task_id)
        if task is None:
            raise ValueError("agent_task missing")
        info = task.input_json or {}
        path = info.get("path")
        entity = info.get("entity")
        account_id = info.get("account_id")
        shop_id = info.get("shop_id")
        if not path or entity is None:
            raise ValueError("import task input_json missing path/entity")
        if not os.path.exists(path):
            raise ValueError(f"upload file missing: {path}")

        # orders：写入前先全量校验同 order_no 相邻（非相邻整文件失败、零脏写）。
        if entity == "orders":
            _prescan_orders(path)

        # reviews 断点续跑：从上次已提交 processed_rows 续；其它从头重跑（幂等）。
        prev = (task.output_json or {}).get("progress") or {}
        resume_offset = int(prev.get("processed_rows", 0)) if entity == "reviews" else 0
        inserted0 = int(prev.get("inserted", 0)) if resume_offset else 0
        updated0 = int(prev.get("updated", 0)) if resume_offset else 0

        err = c.RowErrorCollector(settings.import_max_row_errors)
        if resume_offset:  # 续跑：把已累计的错误带上，最终汇总不丢
            err.count = int(prev.get("error_count", 0))
            err.errors = list(prev.get("errors", []))[: err.cap]

        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb.worksheets[0]
            rows_iter = ws.iter_rows(values_only=True)
            header = next(rows_iter, None)
            if header is None:
                raise ValueError("empty sheet (no header row)")
            mapping = c.map_header(header)

            mod = _ROW_ENTITIES.get(entity)
            required = mod.REQUIRED if mod is not None else orders_imp.REQUIRED
            miss = c.missing_columns(mapping, required)
            if miss:
                raise ValueError(f"missing required columns: {miss}")

            if entity == "orders":
                summary = _drive_orders(
                    db, agent_task_id, rows_iter, mapping, shop_id, err, inserted0, updated0
                )
            else:
                summary = _drive_rows(
                    db, agent_task_id, rows_iter, mapping, mod, account_id, shop_id,
                    err, resume_offset, inserted0, updated0,
                )
        finally:
            wb.close()

        # 与逐块进度同结构（output_json.progress），便于 GET 与续跑读取一致。
        return {"summary": {"progress": summary}, "path": path}
    finally:
        db.close()


def _success(db, agent_task_id: int, result: dict, duration_ms: int) -> None:
    ok = svc.mark_succeeded(
        db, agent_task_id, output_json=result["summary"], duration_ms=duration_ms
    )
    # 仅在确实标记成功后删上传文件；失败/被接管则保留（供 retry 续跑/重跑）。
    if ok and result.get("path"):
        try:
            os.remove(result["path"])
        except OSError:
            pass


@celery_app.task(bind=True, name="app.tasks.imports.run_import", max_retries=100)
def run_import_task(self, agent_task_id: int, payload: dict | None = None):
    run_lifecycle(
        self,
        agent_task_id,
        lambda _p: _run(agent_task_id),
        payload,
        success_handler=_success,
    )
    return {"agent_task_id": agent_task_id}
